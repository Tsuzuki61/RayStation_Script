"""use CPython 2.7"""
from connect import *
import numpy
import os
import re
import openpyxl as px
import sys

sys.path.append(os.path.join(os.path.dirname(__file__), '../common_processing'))
if __name__ == '__main__':
    # for test
    sys.path.append(r'M:\Script\create file\common_processing')
from CommonModules import create_new_patient_folder


class DoseData:
    def __init__(self):
        self.Value = 0.0
        self.perturbed_property = PerturbedDoseProperty()
        self.evaluate_clinical_goal = False

    def setDose(self, value, perturbed_dose_property):
        self.Value = value
        self.perturbed_property = perturbed_dose_property

    def __lt__(self, other):
        """This method defines the sort condition"""
        if self.perturbed_property.BeamSetName < other.perturbed_property.BeamSetName:
            return True
        elif self.perturbed_property.BeamSetName == other.perturbed_property.BeamSetName:
            if self.perturbed_property.density == other.perturbed_property.density == 0.0:
                if self.perturbed_property.x == self.perturbed_property.y == self.perturbed_property.z == 0:
                    return True
                elif other.perturbed_property.x == other.perturbed_property.y == other.perturbed_property.z == 0:
                    return False
                elif self.perturbed_property.x > other.perturbed_property.x:
                    return True
                elif self.perturbed_property.x < other.perturbed_property.x:
                    return False
                elif self.perturbed_property.y > other.perturbed_property.y:
                    return True
                elif self.perturbed_property.y < other.perturbed_property.y:
                    return False
                elif self.perturbed_property.z > other.perturbed_property.z:
                    return True
                elif self.perturbed_property.z < other.perturbed_property.z:
                    return False
                else:
                    return True
            elif self.perturbed_property.density == 0.0 and other.perturbed_property.density != 0.0:
                return True
            elif other.perturbed_property.density == 0.0 and self.perturbed_property.density != 0.0:
                return False
            elif self.perturbed_property.density > other.perturbed_property.density:
                return True
            elif self.perturbed_property.density < other.perturbed_property.density:
                return False
            elif self.perturbed_property.density == other.perturbed_property.density:
                if self.perturbed_property.x == self.perturbed_property.y == self.perturbed_property.z == 0:
                    return True
                elif other.perturbed_property.x == other.perturbed_property.y == other.perturbed_property.z == 0:
                    return False
                elif self.perturbed_property.x > other.perturbed_property.x:
                    return True
                elif self.perturbed_property.x < other.perturbed_property.x:
                    return False
                elif self.perturbed_property.y > other.perturbed_property.y:
                    return True
                elif self.perturbed_property.y < other.perturbed_property.y:
                    return False
                elif self.perturbed_property.z > other.perturbed_property.z:
                    return True
                elif self.perturbed_property.z < other.perturbed_property.z:
                    return False
                else:
                    return True


class PerturbedDoseProperty:
    def __init__(self):
        self.BeamSetName = ''
        self.x = 0.0
        self.y = 0.0
        self.z = 0.0
        self.density = 0.0


class ClinicalGoalSetting:
    def __init__(self):
        self.AcceptanceLevel = 0.0
        self.GoalCriteria = ''
        self.ParameterValue = 0.0
        self.Type = ''
        self.DoseData = []  # DoseDataList

    def DoseDataSort(self):
        self.DoseData = sorted(self.DoseData)

    def getBeamSetDoses(self, BeamSetName):
        rtnDoseData = [tmpDoseData for tmpDoseData in self.DoseData if
                       tmpDoseData.perturbed_property.BeamSetName == BeamSetName]
        return rtnDoseData

    def getMaxDoseinBeamSet(self, BeamSetName):
        tmpDoseData = [tmpDoseData.dose for tmpDoseData in self.DoseData if
                       tmpDoseData.perturbed_property.BeamSetName == BeamSetName]
        return max(tmpDoseData)


class ClinicalGoals:
    def __init__(self):
        self.ClinicalGoalDict = {}

    def setClinicalGoalSetting(self, ROIName, ClinicalGoalSetting):
        if not ROIName in self.ClinicalGoalDict.keys():
            tmpList = [ClinicalGoalSetting]
            self.ClinicalGoalDict.setdefault(ROIName, tmpList)
        else:
            self.ClinicalGoalDict[ROIName].append(ClinicalGoalSetting)


def export_perturbed_dose_in_clinical_goal():
    case = get_current("Case")
    plan = get_current("Plan")
    planningCT_Name = plan.TreatmentCourse.TotalDose.OnDensity.FromExamination.Name
    clinical_goals = ClinicalGoals()
    perturbed_dose_property_list = []
    patient = get_current("Patient")
    ID = patient.PatientID
    Pt_name = patient.Name
    prescription_dose = 0
    for beam_set in plan.BeamSets:
        prescription_dose += beam_set.Prescription.PrimaryDosePrescription.DoseValue
    fractions = len(plan.TreatmentCourse.TreatmentFractions)
    prescription_str = '{}GyRBE/{}Fr'.format(prescription_dose / 100, fractions)

    file_path = create_new_patient_folder()

    os.chdir("M:\\")
    os.chdir(os.getcwd() + file_path)
    # Create xlsx file

    tmpC = 6
    merge_roi_name = ''
    merge_start_c = 0
    wb = px.load_workbook(os.path.join(os.path.dirname(__file__), '../Data/Perturbed_Dose_Clinical_Goals.xlsx'))
    ws = wb['OAR_Template']
    ws.cell(row=1, column=3).value = ID
    ws.cell(row=2, column=3).value = Pt_name
    ws.cell(row=3, column=3).value = prescription_str
    ws.cell(row=4, column=3).value = plan.Name
    EvaluationFunctions_list = sorted(plan.TreatmentCourse.EvaluationSetup.EvaluationFunctions,
                                      key=lambda x: x.ForRegionOfInterest.Name)

    for eFunction in EvaluationFunctions_list:

        tmpClinicalGoalSetting = ClinicalGoalSetting()
        tmpClinicalGoalSetting.AcceptanceLevel = eFunction.PlanningGoal.AcceptanceLevel
        tmpClinicalGoalSetting.GoalCriteria = eFunction.PlanningGoal.GoalCriteria
        tmpClinicalGoalSetting.ParameterValue = eFunction.PlanningGoal.ParameterValue
        tmpClinicalGoalSetting.Type = eFunction.PlanningGoal.Type
        for dose_on_examination in case.TreatmentDelivery.FractionEvaluations[0].DoseOnExaminations:
            if re.match(planningCT_Name, dose_on_examination.OnExamination.Name):
                dose_evaluation_list = dose_on_examination.DoseEvaluations
                break
        # None perturbe dose is NG.
        # for BeamSet in plan.BeamSets:
        #     tmpPerturbedDoseProperty = PerturbedDoseProperty()
        #     tmpPerturbedDoseProperty.BeamSetName = BeamSet.DicomPlanLabel
        #     tmpPerturbedDoseProperty.x = 0.0
        #     tmpPerturbedDoseProperty.y = 0.0
        #     tmpPerturbedDoseProperty.z = 0.0
        #     tmpPerturbedDoseProperty.density = 0.0
        #
        #     # tmpFraction = BeamSet.FractionationPattern.NumberOfFractions
        #
        #     # tmpValue = eFunction.GetClinicalGoalValueForAccumulatedDose(FromFractionNumber=1,
        #     #                                                             ToFractionNumber=tmpFraction,
        #     #                                                             DoseScaling='Actual',
        #     #                                                             DeliveredDose=False)
        #     # tmpEvaluateClinicalGoal = eFunction.EvaluateClinicalGoalForAccumulatedDose(FromFractionNumber=1,
        #     #                                                                            ToFractionNumber=tmpFraction,
        #     #                                                                            DoseScaling='Actual',
        #     #                                                                            DeliveredDose=False)
        #     tmpValue = eFunction.GetClinicalGoalValueForEvaluationDose(DoseDistribution=BeamSet.FractionDose,
        #                                                                ScaleFractionDoseToBeamSet=True)
        #     tmpEvaluateClinicalGoal = eFunction.EvaluateClinicalGoalForEvaluationDose(
        #         DoseDistribution=BeamSet.FractionDose,
        #         ScaleFractionDoseToBeamSet=True)
        #     tmpDoseData = DoseData()
        #     tmpDoseData.setDose(tmpValue, tmpPerturbedDoseProperty)
        #     tmpDoseData.evaluate_clinical_goal = tmpEvaluateClinicalGoal
        #     tmpClinicalGoalSetting.DoseData.append(tmpDoseData)
        for dose_evaluation in dose_evaluation_list:
            if dose_evaluation.PerturbedDoseProperties is not None:
                tmpPerturbedDoseProperty = PerturbedDoseProperty()
                tmpPerturbedDoseProperty.BeamSetName = dose_evaluation.ForBeamSet.DicomPlanLabel
                tmpPerturbedDoseProperty.x = -dose_evaluation.PerturbedDoseProperties.IsoCenterShift.x
                tmpPerturbedDoseProperty.y = -dose_evaluation.PerturbedDoseProperties.IsoCenterShift.z
                tmpPerturbedDoseProperty.z = dose_evaluation.PerturbedDoseProperties.IsoCenterShift.y  # Ver.10A
                tmpPerturbedDoseProperty.density = dose_evaluation.PerturbedDoseProperties.RelativeDensityShift

                tmpValue = eFunction.GetClinicalGoalValueForEvaluationDose(DoseDistribution=dose_evaluation,
                                                                           ScaleFractionDoseToBeamSet=True)
                tmpEvaluateClinicalGoal = eFunction.EvaluateClinicalGoalForEvaluationDose(
                    DoseDistribution=dose_evaluation, ScaleFractionDoseToBeamSet=True)
                tmpDoseData = DoseData()
                tmpDoseData.setDose(tmpValue, tmpPerturbedDoseProperty)
                tmpDoseData.evaluate_clinical_goal = tmpEvaluateClinicalGoal
                tmpClinicalGoalSetting.DoseData.append(tmpDoseData)
        tmpRoi = eFunction.ForRegionOfInterest.Name
        tmpRoiColor = eFunction.ForRegionOfInterest.Color
        tmpRoiHTMLColor = '{R:02X}{G:02X}{B:02X}'.format(R=tmpRoiColor.R, G=tmpRoiColor.G, B=tmpRoiColor.B)
        tmpClinicalGoalSetting.DoseDataSort()

        # xlsx

        ws.cell(5, tmpC).value = tmpRoi
        ws.cell(5, tmpC).alignment = px.styles.alignment.Alignment(horizontal='center', vertical='center')
        ws.cell(5, tmpC).fill = px.styles.fills.PatternFill(patternType='solid', fgColor=tmpRoiHTMLColor)

        if tmpC == 6:
            merge_start_c = tmpC
            merge_roi_name = tmpRoi
        else:
            if not ws.cell(5, tmpC).value == merge_roi_name:
                ws.merge_cells(start_row=5, start_column=merge_start_c, end_row=5, end_column=tmpC - 1)
                merge_start_c = tmpC
                merge_roi_name = tmpRoi
        if re.match('^(Dose).*', tmpClinicalGoalSetting.Type):
            if tmpClinicalGoalSetting.ParameterValue == 0:
                ws.cell(6, tmpC).value = 'Maximum absolute dose'
            elif re.match('.*Absolute.*', tmpClinicalGoalSetting.Type):
                ws.cell(6, tmpC).value = 'Absolute dose at {}cc'.format(tmpClinicalGoalSetting.ParameterValue)
            else:
                ws.cell(6, tmpC).value = 'Absolute dose at {}%'.format(tmpClinicalGoalSetting.ParameterValue * 100)
            if tmpClinicalGoalSetting.GoalCriteria == 'AtMost':
                ws.cell(7, tmpC).value = '<{}cGy'.format(tmpClinicalGoalSetting.AcceptanceLevel)
            elif tmpClinicalGoalSetting.GoalCriteria == 'AtLeast':
                ws.cell(7, tmpC).value = '{}cGy<'.format(tmpClinicalGoalSetting.AcceptanceLevel)
        elif re.match('.*Volume.*', tmpClinicalGoalSetting.Type):
            if re.match('.*Absolute.*', tmpClinicalGoalSetting.Type):
                ws.cell(6, tmpC).value = 'Absolute volume at {}cGy'.format(tmpClinicalGoalSetting.ParameterValue)
                if tmpClinicalGoalSetting.GoalCriteria == 'AtMost':
                    ws.cell(7, tmpC).value = '<{}cc'.format(tmpClinicalGoalSetting.AcceptanceLevel)
                elif tmpClinicalGoalSetting.GoalCriteria == 'AtLeast':
                    ws.cell(7, tmpC).value = '{}cc<'.format(tmpClinicalGoalSetting.AcceptanceLevel)
            else:
                ws.cell(6, tmpC).value = 'Relative volume at {}cGy'.format(tmpClinicalGoalSetting.ParameterValue)
                if tmpClinicalGoalSetting.GoalCriteria == 'AtMost':
                    ws.cell(7, tmpC).value = '<{}%'.format(tmpClinicalGoalSetting.AcceptanceLevel * 100)
                elif tmpClinicalGoalSetting.GoalCriteria == 'AtLeast':
                    ws.cell(7, tmpC).value = '{}%<'.format(tmpClinicalGoalSetting.AcceptanceLevel * 100)
        elif re.match('AverageDose', tmpClinicalGoalSetting.Type):
            ws.cell(6, tmpC).value = 'Average dose'
            if tmpClinicalGoalSetting.GoalCriteria == 'AtMost':
                ws.cell(7, tmpC).value = '<{}cGy'.format(tmpClinicalGoalSetting.AcceptanceLevel)
            elif tmpClinicalGoalSetting.GoalCriteria == 'AtLeast':
                ws.cell(7, tmpC).value = '{}cGy<'.format(tmpClinicalGoalSetting.AcceptanceLevel)
        tmpR = 8
        for dose_data in tmpClinicalGoalSetting.DoseData:
            ws.cell(tmpR, tmpC).value = dose_data.Value
            if not dose_data.evaluate_clinical_goal and not numpy.isnan(dose_data.Value):
                ws.cell(tmpR, tmpC).fill = px.styles.fills.PatternFill(patternType='solid', fgColor='ff0000')
            if tmpC == 6:
                ws.cell(tmpR, 2).value = dose_data.perturbed_property.density
                ws.cell(tmpR, 3).value = dose_data.perturbed_property.x
                ws.cell(tmpR, 4).value = dose_data.perturbed_property.y
                ws.cell(tmpR, 5).value = dose_data.perturbed_property.z
            tmpR += 1
        tmpC += 1

        if len(perturbed_dose_property_list) == 0:
            tmpBeamSetDoses = tmpClinicalGoalSetting.getBeamSetDoses(plan.BeamSets[0].DicomPlanLabel)
            for dose in tmpBeamSetDoses:
                perturbed_dose_property_list.append(dose.perturbed_property)

        # clinical_goals.setClinicalGoalSetting(tmpRoi, tmpClinicalGoalSetting)

    file_name = "Perturbed_Dose_Clinical_Goals_Report_{0}_{1}.xlsx".format(Pt_name, plan.Name)
    # test
    # wb.save(r"DirectoryPath")
    wb.save(file_name)


if __name__ == '__main__':
    export_perturbed_dose_in_clinical_goal()

