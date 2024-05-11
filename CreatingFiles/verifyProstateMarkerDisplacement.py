import os
import pandas as pd
from openpyxl import load_workbook
import re


def verifyProstateMarkerDisplacement(isTest=False):
    folder_path = r"DirectoryPath"
    sheet_name = 'Veri評価'
    data = []
    for filename in os.listdir(folder_path):
        match = re.search(r'\d{8}', filename)
        if match:
            patient_id = match.group()
        else:
            patient_id = None
        file_path = os.path.join(folder_path, filename)
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            target_cells = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value == 'Δマーク・骨' and ws.cell(row=row_idx - 2, column=col_idx - 4).value is not None:
                        target_cells.append((row_idx, col_idx))
            for row_idx, col_idx in target_cells:
                row_data = []
                # append patient ID
                row_data.append(patient_id)
                exam_date = ws.cell(row=row_idx - 2, column=col_idx - 4).value
                # append Vri CT date
                row_data.append(exam_date)
                for i in range(3):
                    cell_value = ws.cell(row=row_idx, column=col_idx + i + 1).value
                    # append X,Y,Z(R-L,S-I,A-P)data
                    row_data.append(cell_value)
                data.append(row_data)
        wb.close()
        if isTest:
            break
    df = pd.DataFrame(data, columns=['PatientID', 'VeriCT_date', 'R-L', 'I-S', 'P-A'])
    df['VeriCT_date'] = pd.to_datetime(df['VeriCT_date'], origin='1899-12-30', unit='D')
    df.to_excel(r'M:\Script\create file\CreatingFiles\veriCT_all_data.xlsx', index=False)
    df[(abs(df['R-L']) > 0.5) | (abs(df['I-S']) > 0.5) | (abs(df['P-A']) > 0.5)].to_excel(
        r'M:\Script\create file\CreatingFiles\veriCT_over_5mm_data.xlsx', index=False)


if __name__ == '__main__':
    verifyProstateMarkerDisplacement()
