# Script recorded 26 Dec 2018

#   RayStation version: 6.2.0.7
#   Selected patient: ...
# use Cpython
from connect import *
import openpyxl as px
from openpyxl.styles import borders, alignment
from decimal import *
import os
import sys

sys.path.append(os.path.join(os.path.dirname(__file__), '../common_processing'))
from CommonModules import create_new_patient_folder


def collation_CSV_create():
    """This function outputs a CSV file for check the dose before each irradiation"""
    plan = get_current("Plan")
    patient = get_current("Patient")
    Pt_name = patient.Name
    ID = patient.PatientID
# -----Retrieve BeamSet data-----
    MUdataDic = {}
    for BeamSet in plan.BeamSets:
        BeamSetDataDic = {}
        for Beam in BeamSet.Beams:
            BeamDataDic = {}
            BeamDataDic.setdefault('BeamName', Beam.Name)
            BeamDataDic.setdefault('MU', Beam.BeamMU)
            BeamSetDataDic.setdefault(Beam.Name, BeamDataDic)
        MUdataDic.setdefault(BeamSet.DicomPlanLabel, BeamSetDataDic)

    BeamSet_list = [BeamSet.DicomPlanLabel for BeamSet in plan.BeamSets]

    file_path = create_new_patient_folder()

    os.chdir("M:\\")
    os.chdir(os.getcwd() + file_path)

    file_name = "Collation_by_reading_out({}).csv".format(plan.Name)

    # write to the file
    with open(file_name, "w") as file:
        file.write("PatientName,{1}\nPatientID,{0}\n".format(ID, Pt_name))
        file.write('\nPlan Name,{}\n'.format(plan.Name))
        file.write("Beam Set,")
        for BeamSetName in BeamSet_list:
            file.write(BeamSetName)
            for BeamName in MUdataDic[BeamSetName].keys():
                file.write(',')
        file.write('\n')
        file.write('Beam,')
        for BeamSetName in BeamSet_list:
            for BeamName in MUdataDic[BeamSetName].keys():
                file.write(BeamName + ',')
        file.write('\n')
        file.write('MU,')
        for BeamSetName in BeamSet_list:
            for BeamName in MUdataDic[BeamSetName].keys():
                file.write(str(Decimal(str(MUdataDic[BeamSetName][BeamName]['MU'])).quantize(Decimal('0.1'),
                                                                                             rounding=ROUND_HALF_UP)) + ',')
        file.write('\n')
        file.write('Preset,')
        for BeamSetName in BeamSet_list:
            for BeamName in MUdataDic[BeamSetName].keys():
                file.write(str(Decimal(str((MUdataDic[BeamSetName][BeamName]['MU'] / 10) * 1.1)).quantize(Decimal('0'),
                                                                                                          rounding=ROUND_HALF_UP)) + ',')


def collation_xlsx_create():
    """This function outputs a CSV file for check the dose before each irradiation"""
    plan = get_current("Plan")
    patient = get_current("Patient")
    Pt_name = patient.Name
    ID = patient.PatientID

    def cell_border(cell):
        cell.border = borders.Border(top=borders.Side(style=borders.BORDER_THIN, color='000000'),
                                     bottom=borders.Side(style=borders.BORDER_THIN, color='000000'),
                                     right=borders.Side(style=borders.BORDER_THIN, color='000000'),
                                     left=borders.Side(style=borders.BORDER_THIN, color='000000'))

    def cell_alignment(cell):
        cell.alignment = alignment.Alignment(horizontal='center', vertical='center', wrap_text=False)

# -----Retrieve BeamSet data-----
    MUdataDic = {}
    for BeamSet in plan.BeamSets:
        BeamSetDataDic = {}
        for Beam in BeamSet.Beams:
            BeamDataDic = {}
            BeamDataDic.setdefault('BeamName', Beam.Name)
            BeamDataDic.setdefault('MU', Beam.BeamMU)
            BeamDataDic.setdefault('Gantry', Beam.GantryAngle)
            BeamDataDic.setdefault('Couch', Beam.CouchRotationAngle)
            BeamSetDataDic.setdefault(Beam.Name, BeamDataDic)
        MUdataDic.setdefault(BeamSet.DicomPlanLabel, BeamSetDataDic)

    BeamSet_list = [BeamSet.DicomPlanLabel for BeamSet in plan.BeamSets]
    BeamSet_list.sort()

    file_path = create_new_patient_folder()

    os.chdir("M:\\")
    os.chdir(os.getcwd() + file_path)

    file_name = "Collation_by_reading_out({}).xlsx".format(plan.Name)
    wb = px.Workbook()
    ws = wb.active

# -----write to the file-----
    ws.cell(1, 1).value = 'Patient Name'
    ws.cell(1, 2).value = Pt_name
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
    ws.cell(2, 1).value = 'Patient ID'
    ws.cell(2, 2).value = ID

    ws.cell(4, 1).value = 'Plan Name'
    cell_border(ws.cell(4, 1))
    ws.cell(4, 2).value = plan.Name
    cell_alignment(ws.cell(4, 2))
    cell_border(ws.cell(4, 2))
    ws.cell(5, 1).value = 'Beam Set'
    cell_border(ws.cell(5, 1))
    ws.cell(6, 1).value = 'Gantry'
    cell_border(ws.cell(6, 1))
    ws.cell(7, 1).value = 'Couch'
    cell_border(ws.cell(7, 1))
    ws.cell(8, 1).value = 'Beam'
    cell_border(ws.cell(8, 1))
    ws.cell(9, 1).value = 'MU'
    cell_border(ws.cell(9, 1))
    ws.cell(10, 1).value = 'Preset'
    cell_border(ws.cell(10, 1))
    tmpC = 2
    merge_start_c = tmpC
# -----write beam set data to the file-----
    for BeamSetName in BeamSet_list:
        ws.cell(5, tmpC).value = BeamSetName
        cell_border(ws.cell(5, tmpC))
        cell_alignment(ws.cell(5, tmpC))
        tmp_beam_name_list = MUdataDic[BeamSetName].keys()
        tmp_beam_name_list.sort()
        for BeamName in tmp_beam_name_list:
            ws.cell(6, tmpC).value = MUdataDic[BeamSetName][BeamName]['Gantry']
            cell_border(ws.cell(6, tmpC))
            cell_alignment(ws.cell(6, tmpC))
            ws.cell(7, tmpC).value = MUdataDic[BeamSetName][BeamName]['Couch']
            cell_border(ws.cell(7, tmpC))
            cell_alignment(ws.cell(7, tmpC))
            ws.cell(8, tmpC).value = BeamName
            cell_border(ws.cell(8, tmpC))
            cell_alignment(ws.cell(8, tmpC))

            ws.cell(9, tmpC).value = str(Decimal(str(MUdataDic[BeamSetName][BeamName]['MU'])).quantize(Decimal('0.1'),
                                                                                                       rounding=ROUND_HALF_UP))
            cell_border(ws.cell(9, tmpC))
            cell_alignment(ws.cell(9, tmpC))
            ws.cell(10, tmpC).value = str(
                Decimal(str((MUdataDic[BeamSetName][BeamName]['MU'] / 10) * 1.1)).quantize(Decimal('0'),
                                                                                           rounding=ROUND_HALF_UP))
            cell_border(ws.cell(10, tmpC))
            cell_alignment(ws.cell(10, tmpC))

            tmpC += 1
        if merge_start_c < tmpC - 1:
            ws.merge_cells(start_row=5, start_column=merge_start_c, end_row=5, end_column=tmpC - 1)
            merge_start_c = tmpC
        else:
            merge_start_c = tmpC
    if merge_start_c < tmpC - 1:
        ws.merge_cells(start_row=5, start_column=merge_start_c, end_row=5, end_column=tmpC - 1)

    tmp_merge_start_gantry = 2
    tmp_merge_start_couch = 2
    tmp_merge_end_gantry = 2
    tmp_merge_end_couch = 2
    for i in range(2, tmpC):
        if ws.cell(6, i).value == ws.cell(6, i + 1).value:
            tmp_merge_end_gantry = i + 1
        elif tmp_merge_start_gantry < tmp_merge_end_gantry:
            ws.merge_cells(start_row=6, start_column=tmp_merge_start_gantry, end_row=6, end_column=tmp_merge_end_gantry)
            tmp_merge_start_gantry = i + 1
        else:
            tmp_merge_start_gantry = i + 1
        if ws.cell(7, i).value == ws.cell(7, i + 1).value:
            tmp_merge_end_couch = i + 1
        elif tmp_merge_start_couch < tmp_merge_end_couch:
            ws.merge_cells(start_row=7, start_column=tmp_merge_start_couch, end_row=7, end_column=tmp_merge_end_couch)
            tmp_merge_start_couch = i + 1
        else:
            tmp_merge_start_couch = i + 1

    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=tmpC - 1)

    # Adjust the width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[px.utils.get_column_letter(column)].width = adjusted_width

    wb.save(file_name)


# test
if __name__ == '__main__':
    collation_xlsx_create()
