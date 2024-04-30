#
# 1) Change CTV name and volume list on Initial parameter.
# 2) Create perturbed dose.
# 3) Script execute.
# use CPython
#

from connect import *
import openpyxl as px
from openpyxl.styles import borders
import os
import sys

sys.path.append(os.path.join(os.path.dirname(__file__), '../common_processing'))
from CommonModules import create_new_patient_folder


class PerturbedData():
    def __init__(self):
        self.DoseData_info = []

    def getBeamSetData(self, BeamSetName):
        tmpList = [PDP for PDP in self.DoseData_info if PDP.BeamSetName == BeamSetName]
        return tmpList

    def getPerturbedPropertyAtMaxDose(self, BeamSetName, Dx):
        tmpList = [PDP for PDP in self.DoseData_info if PDP.BeamSetName == BeamSetName]
        tmpList.sort(key=lambda x: getattr(x.DoseData, Dx), reverse=True)
        return tmpList[0]

    def getPerturbedPropertyAtMinDose(self, BeamSetName, Dx):
        tmpList = [PDP for PDP in self.DoseData_info if PDP.BeamSetName == BeamSetName]
        tmpList.sort(key=lambda x: getattr(x.DoseData, Dx))
        return tmpList[0]

    def getNonPerturbedDoseData(self, BeamSetName, Dx):
        tmpList = [PDP for PDP in self.DoseData_info if PDP.BeamSetName == BeamSetName]
        for perturbed_dose_property in tmpList:
            if perturbed_dose_property.x == 0 and perturbed_dose_property.y == 0 and perturbed_dose_property.z == 0 and perturbed_dose_property.density == 0:
                return getattr(perturbed_dose_property.DoseData, Dx)

    def DoseData_Info_Sort(self):
        self.DoseData_info = sorted(self.DoseData_info)

    def getPerturbedPropertyAtMinDoseFarFromBladder(self, BeamSetName, Dx):
        tmpList = [PDP for PDP in self.DoseData_info if PDP.BeamSetName == BeamSetName and PDP.y < 0 and PDP.z < 0]
        tmpList.sort(key=lambda x: getattr(x.DoseData, Dx))
        return tmpList[0]

    def getPerturbedPropertyAtMaxDoseFarFromBladder(self, BeamSetName, Dx):
        tmpList = [PDP for PDP in self.DoseData_info if PDP.BeamSetName == BeamSetName and PDP.y < 0 and PDP.z < 0]
        tmpList.sort(key=lambda x: getattr(x.DoseData, Dx), reverse=True)
        return tmpList[0]

    def getPerturbedPropertyAtMinDoseFarFromRectum(self, BeamSetName, Dx):
        tmpList = [PDP for PDP in self.DoseData_info if PDP.BeamSetName == BeamSetName and PDP.y > 0 and PDP.z > 0]
        tmpList.sort(key=lambda x: getattr(x.DoseData, Dx))
        return tmpList[0]

    def getPerturbedPropertyAtMaxDoseFarFromRectum(self, BeamSetName, Dx):
        tmpList = [PDP for PDP in self.DoseData_info if PDP.BeamSetName == BeamSetName and PDP.y > 0 and PDP.z > 0]
        tmpList.sort(key=lambda x: getattr(x.DoseData, Dx), reverse=True)
        return tmpList[0]


class DoseData():
    def __init__(self):
        pass

    def setDx(self, VolumeList, DoseList):
        for volume, dose in zip(VolumeList, DoseList):
            setattr(self, "D" + str(int(volume * 100)), dose)

    def getDict(self):
        return self.__dict__

    def setMax(self, MaxDose):
        setattr(self, 'MaxDose', MaxDose)

    def setMin(self, MinDose):
        setattr(self, 'MinDose', MinDose)

    def setAverage(self, AverageDose):
        setattr(self, 'AverageDose', AverageDose)


class PerturbedDoseProperty():
    def __init__(self):
        self.BeamSetName = ''
        self.x = 0.0
        self.y = 0.0
        self.z = 0.0
        self.density = 0.0
        self.DoseData = DoseData()

    def __lt__(self, other):
        """This method defines the sort condition"""
        if self.BeamSetName < other.BeamSetName:
            return True
        elif self.BeamSetName == other.BeamSetName:
            if self.density == other.density == 0.0:
                if self.x == self.y == self.z == 0:
                    return True
                elif other.x == other.y == other.z == 0:
                    return False
                elif self.x > other.x:
                    return True
                elif self.x < other.x:
                    return False
                elif self.y > other.y:
                    return True
                elif self.y < other.y:
                    return False
                elif self.z > other.z:
                    return True
                elif self.z < other.z:
                    return False
                else:
                    return True
            elif self.density == 0.0 and other.density != 0.0:
                return True
            elif other.density == 0.0 and self.density != 0.0:
                return False
            elif self.density > other.density:
                return True
            elif self.density < other.density:
                return False
            elif self.density == other.density:
                if self.x == self.y == self.z == 0:
                    return True
                elif other.x == other.y == other.z == 0:
                    return False
                elif self.x > other.x:
                    return True
                elif self.x < other.x:
                    return False
                elif self.y > other.y:
                    return True
                elif self.y < other.y:
                    return False
                elif self.z > other.z:
                    return True
                elif self.z < other.z:
                    return False
                else:
                    return True


def export_perturbed_dose_of_target_roi_xlsx():
    # Initial parameter
    case = get_current("Case")
    plan = get_current("Plan")

    volume_list = [0.99, 0.98, 0.95, 0.5, 0.02, 0.01]  # D99, D98, D95, D50, D2, D1

    bst_name_list = [bst.DicomPlanLabel for bst in plan.BeamSets]
    PerturbedDoseData = PerturbedData()
    roi_name = []
    for bs in plan.BeamSets:
        if not bs.Prescription.PrimaryDosePrescription.OnStructure.Name in roi_name:
            roi_name.append(bs.Prescription.PrimaryDosePrescription.OnStructure.Name)
        tmpPerturbedData = PerturbedDoseProperty()
        tmpPerturbedData.BeamSetName = bs.DicomPlanLabel
        num_of_fraction = bs.FractionationPattern.NumberOfFractions
        dose_data_1fx = bs.FractionDose.GetDoseAtRelativeVolumes(
            RoiName=bs.Prescription.PrimaryDosePrescription.OnStructure.Name,
            RelativeVolumes=volume_list)
        dose_data_total = [dose * num_of_fraction for dose in dose_data_1fx]
        tmpPerturbedData.DoseData.setDx(volume_list, dose_data_total)
        PerturbedDoseData.DoseData_info.append(tmpPerturbedData)
        if 'Prostate' not in plan.Name:
            tmpAverageDose = bs.FractionDose.GetDoseStatistic(
                RoiName=bs.FractionDose.ForBeamSet.Prescription.PrimaryDosePrescription.OnStructure.Name,
                DoseType="Average") * num_of_fraction
            tmpPerturbedData.DoseData.setAverage(tmpAverageDose)

    for DoseOnE in case.TreatmentDelivery.FractionEvaluations[0].DoseOnExaminations:
        for eval_dose in DoseOnE.DoseEvaluations:
            if eval_dose.PerturbedDoseProperties is not None and eval_dose.ForBeamSet.DicomPlanLabel in bst_name_list:
                if plan.BeamSets[eval_dose.ForBeamSet.DicomPlanLabel].Prescription.PrimaryDosePrescription.DoseValue \
                        == eval_dose.ForBeamSet.Prescription.PrimaryDosePrescription.DoseValue:
                    tmpPerturbedData = PerturbedDoseProperty()
                    tmpPerturbedData.BeamSetName = eval_dose.ForBeamSet.DicomPlanLabel
                    tmpPerturbedData.x = -eval_dose.PerturbedDoseProperties.IsoCenterShift.x
                    tmpPerturbedData.y = -eval_dose.PerturbedDoseProperties.IsoCenterShift.z
                    tmpPerturbedData.z = eval_dose.PerturbedDoseProperties.IsoCenterShift.y # Ver.10A
                    tmpPerturbedData.density = eval_dose.PerturbedDoseProperties.RelativeDensityShift

                    num_of_fraction = eval_dose.ForBeamSet.FractionationPattern.NumberOfFractions
                    dose_data_1fx = eval_dose.GetDoseAtRelativeVolumes(
                        RoiName=eval_dose.ForBeamSet.Prescription.PrimaryDosePrescription.OnStructure.Name,
                        RelativeVolumes=volume_list)
                    dose_data_total = [dose * num_of_fraction for dose in dose_data_1fx]

                    tmpPerturbedData.DoseData.setDx(volume_list, dose_data_total)
                    if 'Prostate' not in plan.Name:
                        tmpAverageDose = eval_dose.GetDoseStatistic(
                            RoiName=eval_dose.ForBeamSet.Prescription.PrimaryDosePrescription.OnStructure.Name,
                            DoseType="Average") * num_of_fraction
                        tmpPerturbedData.DoseData.setAverage(tmpAverageDose)

                    # PerturbedDataList append
                    PerturbedDoseData.DoseData_info.append(tmpPerturbedData)

    # create file & folder
    patient = get_current("Patient")
    ID = patient.PatientID
    Pt_name = patient.Name
    prescription_dose = 0.0
    prescription_fraction = 0

    for beam_set in plan.BeamSets:
        prescription_dose += beam_set.Prescription.PrimaryDosePrescription.DoseValue
        prescription_fraction += beam_set.FractionationPattern.NumberOfFractions

    file_path = create_new_patient_folder()

    os.chdir("M:\\")
    os.chdir(os.getcwd() + file_path)

    PerturbedDoseData.DoseData_Info_Sort()

    template_file_name = os.path.join(os.path.dirname(__file__), '../Data/TreatmentPlanning_Evaluation_CTV.xlsx')
    isProstate = False
    if 'Prostate' in plan.Name:
        template_file_name = os.path.join(os.path.dirname(__file__),
                                          '../Data/TreatmentPlanning_Evaluation_CTV_Prostate.xlsx')
        isProstate = True
    file_name = "Perturbed_Dose_for_Excel_Report_{0}_{1}.xlsx".format(Pt_name, plan.Name)

    wb = px.load_workbook(template_file_name)

    ws = wb.worksheets[0]

    ws.cell(1, 2).value = ID
    ws.cell(2, 2).value = Pt_name
    ws.cell(3, 2).value = '{:.1f}GyRBE/{}fr'.format(prescription_dose / 100, prescription_fraction)
    for BeamSetName in bst_name_list:
        prescription = plan.BeamSets[BeamSetName].Prescription.PrimaryDosePrescription.DoseValue
        writing_ws = wb.copy_worksheet(wb.worksheets[0])
        writing_ws.title = BeamSetName
        writing_ws.cell(7, 5).value = BeamSetName
        writing_ws.freeze_panes = 'E10'
        tmpPerturbedDataList = PerturbedDoseData.getBeamSetData(BeamSetName)
        i = 10
        merge_start_row = 10
        for tmpPD in tmpPerturbedDataList:
            writing_ws.cell(i, 1).value = tmpPD.density * 100
            writing_ws.cell(i, 2).value = tmpPD.x
            writing_ws.cell(i, 3).value = tmpPD.y
            writing_ws.cell(i, 4).value = tmpPD.z
            if isProstate:
                KeyList = tmpPD.DoseData.getDict().keys()
                KeyList.sort(reverse=True)
                border_r = 11
            else:
                KeyList = ['D99', 'D98', 'D95', 'AverageDose', 'D50', 'D2', 'D1']
                border_r = 12
            j = 5
            for key in KeyList:
                writing_ws.cell(i, j).value = getattr(tmpPD.DoseData, key)
                j += 1
            if not i == 10:
                if not writing_ws.cell(i, 1).value == writing_ws.cell(i - 1, 1).value:
                    for c in range(1, border_r):
                        writing_ws.cell(i - 1, c).border = borders.Border(
                            bottom=borders.Side(style=borders.BORDER_THIN, color='000000'))
                    writing_ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=i - 1, end_column=1)
                    merge_start_row = i
            i += 1

        for c in range(1, border_r):
            writing_ws.cell(i - 1, c).border = borders.Border(
                bottom=borders.Side(style=borders.BORDER_THIN, color='000000'))
        writing_ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=i - 1, end_column=1)

        summary_start_r = 40
        summary_start_c = 5
        for key in KeyList:
            if len(key) >= 3:
                tmpData = PerturbedDoseData.getPerturbedPropertyAtMinDose(BeamSetName, key)
            else:
                tmpData = PerturbedDoseData.getPerturbedPropertyAtMaxDose(BeamSetName, key)
            writing_ws.cell(summary_start_r, summary_start_c).value = getattr(tmpData.DoseData, key)
            writing_ws.cell(summary_start_r, summary_start_c + 1).value = getattr(tmpData.DoseData,
                                                                                  key) / PerturbedDoseData.getNonPerturbedDoseData(
                BeamSetName, key) * 100
            writing_ws.cell(summary_start_r, summary_start_c + 2).value = tmpData.density * 100
            writing_ws.cell(summary_start_r, summary_start_c + 3).value = tmpData.x
            writing_ws.cell(summary_start_r, summary_start_c + 4).value = tmpData.y
            writing_ws.cell(summary_start_r, summary_start_c + 5).value = tmpData.z
            summary_start_r += 1

        if isProstate:
            # far from bladder
            summary_start_r = 49
            summary_start_c = 5
            for key in KeyList:
                if len(key) >= 3:
                    tmpData = PerturbedDoseData.getPerturbedPropertyAtMinDoseFarFromBladder(BeamSetName, key)
                else:
                    tmpData = PerturbedDoseData.getPerturbedPropertyAtMaxDoseFarFromBladder(BeamSetName, key)
                writing_ws.cell(summary_start_r, summary_start_c).value = getattr(tmpData.DoseData, key)
                writing_ws.cell(summary_start_r, summary_start_c + 1).value = getattr(tmpData.DoseData,
                                                                                      key) / PerturbedDoseData.getNonPerturbedDoseData(
                    BeamSetName, key) * 100
                writing_ws.cell(summary_start_r, summary_start_c + 2).value = tmpData.density * 100
                writing_ws.cell(summary_start_r, summary_start_c + 3).value = tmpData.x
                writing_ws.cell(summary_start_r, summary_start_c + 4).value = tmpData.y
                writing_ws.cell(summary_start_r, summary_start_c + 5).value = tmpData.z
                summary_start_r += 1

            # far from rectum
            summary_start_r = 58
            summary_start_c = 5
            for key in KeyList:
                if len(key) >= 3:
                    tmpData = PerturbedDoseData.getPerturbedPropertyAtMinDoseFarFromRectum(BeamSetName, key)
                else:
                    tmpData = PerturbedDoseData.getPerturbedPropertyAtMaxDoseFarFromRectum(BeamSetName, key)
                writing_ws.cell(summary_start_r, summary_start_c).value = getattr(tmpData.DoseData, key)
                writing_ws.cell(summary_start_r, summary_start_c + 1).value = getattr(tmpData.DoseData,
                                                                                      key) / PerturbedDoseData.getNonPerturbedDoseData(
                    BeamSetName, key) * 100
                writing_ws.cell(summary_start_r, summary_start_c + 2).value = tmpData.density * 100
                writing_ws.cell(summary_start_r, summary_start_c + 3).value = tmpData.x
                writing_ws.cell(summary_start_r, summary_start_c + 4).value = tmpData.y
                writing_ws.cell(summary_start_r, summary_start_c + 5).value = tmpData.z
                summary_start_r += 1
    wb.remove(wb.worksheets[0])
    wb.save(file_name)


### End ###

# test
if __name__ == '__main__':
    export_perturbed_dose_of_target_roi_xlsx()
